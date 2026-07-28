"""Export curated voter records as Excel, CSV or PDF.

The PDF path is the interesting one. ReportLab's built-in fonts (Helvetica
and friends) contain no Tamil glyphs, so a report generated with them renders
every name as a row of empty boxes -- output that looks like a working
feature while being completely unusable. This module therefore *resolves and
verifies* a Tamil-capable font before rendering, and fails with an actionable
message rather than producing tofu.
"""

from __future__ import annotations

import csv
import io
import logging
from datetime import datetime, timezone
from pathlib import Path

from ..config import settings

logger = logging.getLogger(__name__)

# A Tamil consonant. If a font's cmap lacks this, it cannot render the data.
TAMIL_PROBE = 0x0B95

COLUMNS: list[tuple[str, str]] = [
    ("serial", "S.No"),
    ("epic", "EPIC ID"),
    ("name", "Name"),
    ("relation_type", "Relation"),
    ("relation_name", "Relation Name"),
    ("house_number", "House No."),
    ("age", "Age"),
    ("gender", "Gender"),
    ("part_number", "Part"),
]

EXTRA_COLUMNS: list[tuple[str, str]] = [
    ("constituency", "Constituency"),
    ("source_file_name", "Source File"),
    ("page_number", "Page"),
    ("verified", "Verified"),
    ("notes", "Notes"),
    ("created_at", "Created"),
    ("updated_by", "Updated By"),
]

# Where a Tamil-capable font usually lives. `subfont` indexes into a .ttc.
FONT_CANDIDATES: list[tuple[str, int]] = [
    (r"C:\Windows\Fonts\Nirmala.ttc", 0),
    (r"C:\Windows\Fonts\latha.ttf", 0),
    ("/usr/share/fonts/truetype/noto/NotoSansTamil-Regular.ttf", 0),
    ("/usr/share/fonts/truetype/noto/NotoSerifTamil-Regular.ttf", 0),
    ("/usr/share/fonts/truetype/lohit-tamil/Lohit-Tamil.ttf", 0),
    ("/usr/share/fonts/opentype/noto/NotoSansTamil-Regular.otf", 0),
    ("/System/Library/Fonts/Supplemental/Tamil MN.ttc", 0),
]


class ExportError(Exception):
    pass


# ---------------------------------------------------------------------------
# Font resolution
# ---------------------------------------------------------------------------


def _covers_tamil(path: Path, subfont: int) -> bool:
    """True when the font actually has Tamil glyphs.

    Uses fontTools when available. If it is not installed we optimistically
    accept the font rather than blocking the export -- the name-based
    candidate list is a reasonable proxy.
    """
    try:
        from fontTools.ttLib import TTCollection, TTFont
    except ImportError:  # pragma: no cover - optional verification
        return True

    try:
        if path.suffix.lower() == ".ttc":
            font = TTCollection(str(path)).fonts[subfont]
        else:
            font = TTFont(str(path), fontNumber=subfont)
        return any(
            TAMIL_PROBE in table.cmap
            for table in font["cmap"].tables
            if table.isUnicode()
        )
    except Exception as exc:  # noqa: BLE001
        logger.debug("Could not inspect %s: %s", path, exc)
        return False


def resolve_font() -> tuple[Path, int]:
    """Locate a font that can render Tamil, or explain how to supply one."""
    candidates: list[tuple[str, int]] = []
    if settings.pdf_font_path:
        candidates.append((settings.pdf_font_path, 0))
    candidates.extend(FONT_CANDIDATES)

    tried: list[str] = []
    for raw, subfont in candidates:
        path = Path(raw)
        if not path.is_file():
            tried.append(f"{raw} (not found)")
            continue
        if not _covers_tamil(path, subfont):
            tried.append(f"{raw} (no Tamil glyphs)")
            continue
        return path, subfont

    raise ExportError(
        "PDF export needs a font containing Tamil glyphs, and none was found. "
        "Set OCR_PDF_FONT_PATH to a .ttf/.ttc that has them, or install one "
        "(Debian/Ubuntu: `apt-get install fonts-noto-tamil`). Tried: "
        + "; ".join(tried)
    )


_registered: str | None = None


def _register_font() -> str:
    """Register the Tamil font with ReportLab once, returning its name."""
    global _registered
    if _registered:
        return _registered

    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont as RLTTFont

    path, subfont = resolve_font()
    name = "TamilBody"
    try:
        pdfmetrics.registerFont(RLTTFont(name, str(path), subfontIndex=subfont))
    except TypeError:
        # Older ReportLab builds have no subfontIndex parameter.
        pdfmetrics.registerFont(RLTTFont(name, str(path)))
    logger.info("PDF font: %s (subfont %d)", path, subfont)
    _registered = name
    return name


# ---------------------------------------------------------------------------
# Row shaping
# ---------------------------------------------------------------------------


def _columns(include_meta: bool) -> list[tuple[str, str]]:
    return COLUMNS + EXTRA_COLUMNS if include_meta else COLUMNS


def _cell(row, key: str) -> str:
    value = getattr(row, key, "")
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    return str(value)


def build_table(rows, include_meta: bool = False) -> tuple[list[str], list[list[str]]]:
    cols = _columns(include_meta)
    header = [label for _, label in cols]
    data = [[_cell(r, key) for key, _ in cols] for r in rows]
    return header, data


# ---------------------------------------------------------------------------
# Writers
# ---------------------------------------------------------------------------


def to_csv(rows, include_meta: bool = False) -> bytes:
    header, data = build_table(rows, include_meta)
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer)
    writer.writerow(header)
    writer.writerows(data)
    # BOM so Excel detects UTF-8 on a double-click instead of guessing the
    # system codepage and mangling every Tamil name.
    return b"\xef\xbb\xbf" + buffer.getvalue().encode("utf-8")


def to_xlsx(rows, include_meta: bool = False) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header, data = build_table(rows, include_meta)
    wb = Workbook()
    ws = wb.active
    ws.title = "Voters"

    ws.append(header)
    fill = PatternFill("solid", fgColor="4F46E5")
    bold = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = bold
        cell.alignment = Alignment(vertical="center")

    for record in data:
        ws.append(record)

    ws.freeze_panes = "A2"
    if data:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{len(data) + 1}"

    for i, label in enumerate(header, start=1):
        widest = max([len(label)] + [len(r[i - 1]) for r in data[:300]] or [0])
        ws.column_dimensions[get_column_letter(i)].width = min(max(widest + 2, 10), 42)

    meta = wb.create_sheet("Report Info")
    meta.append(["Generated", datetime.now(timezone.utc).isoformat(timespec="seconds")])
    meta.append(["Voter records", len(data)])
    meta.column_dimensions["A"].width = 18
    meta.column_dimensions["B"].width = 52

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def to_pdf(rows, include_meta: bool = False, title: str = "Voter Records") -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    font = _register_font()
    header, data = build_table(rows, include_meta)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TamilTitle", parent=styles["Title"], fontName=font, fontSize=15, spaceAfter=4
    )
    meta_style = ParagraphStyle(
        "TamilMeta", parent=styles["Normal"], fontName=font, fontSize=8,
        textColor=colors.HexColor("#64748B"), alignment=TA_LEFT,
    )
    # Cells are Paragraphs, not raw strings, so long Tamil names wrap inside
    # the column instead of overflowing the page.
    cell_style = ParagraphStyle(
        "TamilCell", parent=styles["Normal"], fontName=font, fontSize=7.5, leading=9.5
    )
    head_style = ParagraphStyle(
        "TamilHead", parent=cell_style, fontSize=7.5, textColor=colors.white
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=10 * mm, bottomMargin=12 * mm,
        title=title, author="OCR Workspace",
    )

    story = [
        Paragraph(title, title_style),
        Paragraph(
            f"{len(data)} record(s) &nbsp;·&nbsp; generated "
            f"{datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
            meta_style,
        ),
        Spacer(1, 6),
    ]

    if not data:
        story.append(Paragraph("No records matched the selection.", cell_style))
    else:
        table_data = [[Paragraph(h, head_style) for h in header]]
        table_data += [[Paragraph(c or "", cell_style) for c in row] for row in data]

        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#F8FAFC")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(table)

    def footer(canvas, doc_):
        canvas.saveState()
        canvas.setFont(font, 7)
        canvas.setFillColor(colors.HexColor("#94A3B8"))
        canvas.drawRightString(
            doc_.pagesize[0] - 10 * mm, 6 * mm, f"Page {canvas.getPageNumber()}"
        )
        canvas.drawString(10 * mm, 6 * mm, title)
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    return buffer.getvalue()


MEDIA_TYPES = {
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "csv": "text/csv; charset=utf-8",
    "pdf": "application/pdf",
}


def export(rows, fmt: str, include_meta: bool = False) -> tuple[bytes, str, str]:
    """Render an export. Returns (content, filename, media_type)."""
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    filename = f"voters-{stamp}.{fmt}"

    if fmt == "csv":
        return to_csv(rows, include_meta), filename, MEDIA_TYPES["csv"]
    if fmt == "xlsx":
        return to_xlsx(rows, include_meta), filename, MEDIA_TYPES["xlsx"]
    if fmt == "pdf":
        return to_pdf(rows, include_meta), filename, MEDIA_TYPES["pdf"]
    raise ExportError(f"Unsupported format: {fmt}")
