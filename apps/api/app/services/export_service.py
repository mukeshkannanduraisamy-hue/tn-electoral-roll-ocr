"""Export extracted records to xlsx / csv / json / txt / markdown.

Three modes, because "export the data" means different things at different
points in a workflow:

``all``    every record, effective values (edit if present, else OCR).
``clean``  only records with zero validation errors -- the subset safe to
           hand downstream without further review.
``audit``  every record with original value, edited value, confidence and
           issues side by side, for checking extraction quality.
"""

from __future__ import annotations

import csv
import io
import json
from datetime import datetime, timezone

from ..schemas.core import (
    ColumnDef,
    ExportFormat,
    ExportMode,
    ExportRequest,
    IssueSeverity,
    Page,
    Record,
)
from ..templates import registry


class ExportError(Exception):
    pass


# ---------------------------------------------------------------------------
# Selection and shaping
# ---------------------------------------------------------------------------


def _record_errors(record: Record) -> int:
    n = 0
    for issue in [*record.issues, *(i for f in record.fields.values() for i in f.issues)]:
        severity = issue.severity if isinstance(issue.severity, str) else issue.severity.value
        if severity == IssueSeverity.ERROR.value:
            n += 1
    return n


def _record_issue_text(record: Record) -> str:
    parts = [i.message for i in record.issues]
    for field in record.fields.values():
        parts.extend(f"{field.key}: {i.message}" for i in field.issues)
    return "; ".join(parts)


def select_records(
    pages: list[Page], request: ExportRequest
) -> list[tuple[Page, Record]]:
    """Apply the request's filters, preserving document order."""
    wanted_pages = set(request.page_ids or [])
    wanted_records = set(request.record_ids or [])
    wanted_files = set(request.file_ids or [])

    selected: list[tuple[Page, Record]] = []
    for page in pages:
        if wanted_files and page.file_id not in wanted_files:
            continue
        if wanted_pages and page.id not in wanted_pages:
            continue
        for record in page.records:
            if wanted_records and record.id not in wanted_records:
                continue
            if request.mode == ExportMode.CLEAN.value and _record_errors(record) > 0:
                continue
            selected.append((page, record))
    return selected


def _columns_for(pages: list[Page]) -> list[ColumnDef]:
    """Union of the columns of every template present, in first-seen order."""
    columns: list[ColumnDef] = []
    seen: set[str] = set()
    for page in pages:
        template = registry.get_or_generic(page.template_id)
        for col in template.columns():
            if col.key not in seen:
                seen.add(col.key)
                columns.append(col)
    return columns


def build_table(
    pages: list[Page], request: ExportRequest
) -> tuple[list[str], list[list[str]]]:
    """Flatten the selection into a header row plus data rows."""
    columns = _columns_for(pages)
    rows_source = select_records(pages, request)
    audit = request.mode == ExportMode.AUDIT.value

    header: list[str] = []
    if request.include_page_numbers:
        header += ["File", "Page", "Row"]

    for col in columns:
        header.append(col.label)
        if audit:
            header.append(f"{col.label} (OCR)")
        if audit or request.include_confidence:
            header.append(f"{col.label} (conf)")

    if audit or request.include_issues:
        header.append("Issues")

    rows: list[list[str]] = []
    for page, record in rows_source:
        row: list[str] = []
        if request.include_page_numbers:
            row += [page.file_id, str(page.page_number), str(record.index + 1)]

        for col in columns:
            field = record.fields.get(col.key)
            row.append(field.value if field else "")
            if audit:
                row.append(field.original_value if field else "")
            if audit or request.include_confidence:
                row.append(f"{field.confidence:.4f}" if field else "")

        if audit or request.include_issues:
            row.append(_record_issue_text(record))

        rows.append(row)

    return header, rows


# ---------------------------------------------------------------------------
# Format writers
# ---------------------------------------------------------------------------


def _to_csv(header: list[str], rows: list[list[str]]) -> bytes:
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, quoting=csv.QUOTE_MINIMAL)
    writer.writerow(header)
    writer.writerows(rows)
    # BOM so Excel opens UTF-8 Tamil correctly on a double-click; without it
    # Excel guesses the system codepage and renders mojibake.
    return b"\xef\xbb\xbf" + buffer.getvalue().encode("utf-8")


def _to_xlsx(header: list[str], rows: list[list[str]], pages: list[Page]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Records"

    ws.append(header)
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    for row in rows:
        ws.append(row)

    # Freeze the header and enable filtering -- this is a review artefact,
    # not just a data dump.
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = (
            f"A1:{get_column_letter(len(header))}{len(rows) + 1}"
        )

    for i, name in enumerate(header, start=1):
        longest = max([len(str(name))] + [len(str(r[i - 1])) for r in rows[:200]] or [0])
        ws.column_dimensions[get_column_letter(i)].width = min(max(longest + 2, 10), 45)

    # Provenance sheet: an export with no record of how it was produced is
    # hard to trust six months later.
    meta = wb.create_sheet("Export Info")
    meta.append(["Generated", datetime.now(timezone.utc).isoformat(timespec="seconds")])
    meta.append(["Records", len(rows)])
    meta.append(["Pages", len(pages)])
    meta.append(["Templates", ", ".join(sorted({p.template_id or "" for p in pages}))])
    meta.column_dimensions["A"].width = 18
    meta.column_dimensions["B"].width = 60

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _to_json(pages: list[Page], request: ExportRequest) -> bytes:
    audit = request.mode == ExportMode.AUDIT.value
    payload = []
    for page, record in select_records(pages, request):
        entry: dict = {
            "file_id": page.file_id,
            "page_number": page.page_number,
            "record_index": record.index,
            "template_id": record.template_id,
            "fields": {},
        }
        for key, field in record.fields.items():
            if audit:
                entry["fields"][key] = {
                    "value": field.value,
                    "original_value": field.original_value,
                    "edited_value": field.edited_value,
                    "suggested_value": field.suggested_value,
                    "confidence": field.confidence,
                    "issues": [i.model_dump() for i in field.issues],
                }
            elif request.include_confidence:
                entry["fields"][key] = {
                    "value": field.value,
                    "confidence": field.confidence,
                }
            else:
                entry["fields"][key] = field.value

        if audit or request.include_issues:
            entry["issues"] = [i.model_dump() for i in record.issues]
        payload.append(entry)

    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


def _to_txt(pages: list[Page], request: ExportRequest) -> bytes:
    lines: list[str] = []
    current_page: str | None = None

    for page, record in select_records(pages, request):
        if request.include_page_numbers and page.id != current_page:
            current_page = page.id
            if lines:
                lines.append("")
            lines.append(f"--- Page {page.page_number} ---")
            if page.header_text:
                lines.append(page.header_text)
            lines.append("")

        parts = [f"{k}: {f.value}" for k, f in record.fields.items() if f.value]
        lines.append(" | ".join(parts))

    return ("\n".join(lines) + "\n").encode("utf-8")


def _to_markdown(header: list[str], rows: list[list[str]]) -> bytes:
    def escape(value: str) -> str:
        return str(value).replace("|", "\\|").replace("\n", " ")

    out = ["| " + " | ".join(escape(h) for h in header) + " |"]
    out.append("| " + " | ".join("---" for _ in header) + " |")
    for row in rows:
        out.append("| " + " | ".join(escape(c) for c in row) + " |")
    return ("\n".join(out) + "\n").encode("utf-8")


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

MEDIA_TYPES = {
    ExportFormat.XLSX.value: (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ),
    ExportFormat.CSV.value: "text/csv; charset=utf-8",
    ExportFormat.JSON.value: "application/json; charset=utf-8",
    ExportFormat.TXT.value: "text/plain; charset=utf-8",
    ExportFormat.MARKDOWN.value: "text/markdown; charset=utf-8",
}


def export(pages: list[Page], request: ExportRequest) -> tuple[bytes, str, str]:
    """Render an export. Returns (content, filename, media_type)."""
    fmt = request.format if isinstance(request.format, str) else request.format.value
    mode = request.mode if isinstance(request.mode, str) else request.mode.value

    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    filename = f"ocr-export-{mode}-{stamp}.{fmt}"

    if fmt in (ExportFormat.CSV.value, ExportFormat.XLSX.value,
               ExportFormat.MARKDOWN.value):
        header, rows = build_table(pages, request)
        if fmt == ExportFormat.CSV.value:
            content = _to_csv(header, rows)
        elif fmt == ExportFormat.XLSX.value:
            content = _to_xlsx(header, rows, pages)
        else:
            content = _to_markdown(header, rows)
    elif fmt == ExportFormat.JSON.value:
        content = _to_json(pages, request)
    elif fmt == ExportFormat.TXT.value:
        content = _to_txt(pages, request)
    else:
        raise ExportError(f"Unsupported export format: {fmt}")

    return content, filename, MEDIA_TYPES.get(fmt, "application/octet-stream")


def preview(pages: list[Page], request: ExportRequest, limit: int = 50) -> dict:
    """Header + first `limit` rows, for the pre-download preview."""
    header, rows = build_table(pages, request)
    return {
        "columns": header,
        "rows": rows[:limit],
        "total_rows": len(rows),
        "truncated": len(rows) > limit,
    }
